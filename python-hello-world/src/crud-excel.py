import uuid
import openpyxl
import os


def map_customer_data(customer: enumerate):
    print(f'Customer: {str(customer)}, Index: {1234}')
    return [
        {f'A{1234}': customer['id']},
        {f'B{1234}': customer['name']},
        {f'C{1234}': customer['email']},
    ]


def modify_excel_file(file_path: str = ''):
    print("Creating and updating excel file")

    # Cria uma instância do workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    cell_content_A1 = ' '.join([
        "Hi, I'm a sheet modified by code in python.",
        "File ./src/crud-excel.py"
    ])

    sheet.merge_cells('A1:C1')
    sheet['A1'] = cell_content_A1

    sheet['A2'] = 'ID'
    sheet['B2'] = 'Name'
    sheet['C2'] = 'E-mail'

    customers = [
        {'id': 1, 'name': 'John Doe', 'email': 'john@teste.com'},
        {'id': 2, 'name': 'Ed Sheeran', 'email': 'ed@teste.com'},
        {'id': 3, 'name': 'Bob Esponja', 'email': 'bob@teste.com'},
        {'id': 4, 'name': 'João Frango', 'email': 'joao@teste.com'},
    ]

    start_row_index = 3

    for customer in customers:
        sheet[f'A{start_row_index}'] = customer['id']
        sheet[f'B{start_row_index}'] = customer['name']
        sheet[f'C{start_row_index}'] = customer['email']
        start_row_index += 1

    # Salva o workbook no arquivo
    workbook.save(file_path)

    print("Excel file created and updated successfully.")


def read_excel_file(file_path: str):
    print(f'Reading excel file: {file_path}')

    # Abre o arquivo Excel usando openpyxl
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Itera sobre as linhas da planilha e imprime o conteúdo
    for row in sheet.iter_rows():
        for cell in row:
            print(cell.value, end='\t')
        print()

    # Não é necessário fechar explicitamente o arquivo, openpyxl cuida disso


def delete_excel_file(file_path: str = ''):
    print("Deleting excel file")
    try:
        os.remove(file_path)
        print(f"File Excel '{file_path}' removed successful.")
    except FileNotFoundError:
        print(f"File Excel '{file_path}' not found.")
    except Exception as e:
        print(f"Error to remove file excel: {e}")


def main():
    id_file = str(uuid.uuid4())
    file_path = f'./temp/{id_file}.xlsx'

    modify_excel_file(file_path)
    # read_excel_file(file_path)
    # delete_excel_file(file_path)


main()
